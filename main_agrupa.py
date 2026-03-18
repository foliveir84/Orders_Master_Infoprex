import streamlit as st
from dotenv import load_dotenv
from datetime import datetime
import os
from sqlalchemy import create_engine
from sqlalchemy.exc import OperationalError
import pandas as pd
from io import BytesIO
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import PyPDF2
import time
import re
import unicodedata


# Carrega as variáveis do ficheiro .env
load_dotenv()


# Obter o URL para ligação a base de dados a partir da variavel de ambiente -> Base de dados projecto Pharmaruptures
DATABASE_URL = os.getenv("DATABASE_URL")
GOOGLE_SHEETS = os.getenv("GOOGLE_SHEETS")


pd.options.display.float_format = '{:.2f}'.format

st.set_page_config(layout='wide')

# Inicializa uploader_key no session_state se ainda não existir
if 'uploader_key' not in st.session_state:
    st.session_state['uploader_key'] = 0


def limpar_designacao(texto):
    """Remove acentos, asteriscos e converte para maiúsculas para ordenação correta e consistência."""
    if not isinstance(texto, str):
        return str(texto)
    # Normaliza unicode (NFD decompõe caracteres) e remove marcas de acentuação (Mn)
    texto_limpo = ''.join(c for c in unicodedata.normalize('NFD', texto)
                          if unicodedata.category(c) != 'Mn')
    # Remove asteriscos
    texto_limpo = texto_limpo.replace('*', '')
    return texto_limpo.strip().title()


# ==========================================Funcoes para obter a dataframe de rupturas do Infarmed========================
@st.cache_data(show_spinner="🔄 A carregar dados da base de dados de Esgotados", ttl=3600)
def obter_base_dados_esgotados(max_wait: int = 30):
    """ Funcao que vai buscar a base de dados do Pharmaruptures 
    com todos os produtos esgotados. Vai agurdar no máximo max_wait segundos.
    Se existir algum erro, vai devolver uma dataframe vazia
    """

    df_esgotados = pd.DataFrame()

    try:
        df_esgotados = pd.read_excel(DATABASE_URL)

        colunas = ['Número de registo',
                   'Nome do medicamento',
                   'Data de início de rutura',
                   'Data prevista para reposição',
                   'TimeDelta',
                   'Data da Consulta',
                   ]
        df_esgotados = df_esgotados[colunas].copy()

        df_esgotados['Número de registo'] = df_esgotados['Número de registo'].astype(
            str)
        df_esgotados['TimeDelta'] = pd.to_numeric(
            df_esgotados['TimeDelta'], errors='coerce')

        # Adicionada variavel para obter a data da Consulta da Base de Dados
        data_consulta = df_esgotados["Data da Consulta"].iloc[0]
        data_consulta = str(data_consulta)[:10]
        # Para adicionar essa informacao ao Stremlit

        return df_esgotados, data_consulta  # ✅ sucesso, retorna DataFrame

    except Exception as e:
        # erro inesperado, regista no log do Streamlit
        st.warning(f"⚠️ Nao foi possível carregar a BD de Esgotados")
        print(e)

    df_esgotados = pd.DataFrame()
    data_consulta = 'Nao Foi possível carregar a INFO'
    return df_esgotados, data_consulta  # Devoolve uma datagrame vazia se falhar


def extract_text_from_pdf(pdf_path):
    pdf_bytes = pdf_path.read()
    pdf_file = BytesIO(pdf_bytes)

    reader = PyPDF2.PdfReader(pdf_file)
    text = ""
    for page_num in range(len(reader.pages)):
        text += reader.pages[page_num].extract_text()
    return text


def extract_data_from_text(text):
    lines = text.split('\n')
    products = []
    codes = []
    quantities = []
    pattern = re.compile(r'^(\d+\,\d{2}€) (\d+) (.*?)(\d{7}) ')

    for line in lines:
        match = pattern.match(line)
        if match:
            price, quantity, product, code = match.groups()
            products.append(product.strip())
            codes.append(int(code.strip()))
            quantities.append(int(quantity.strip()))

    df = pd.DataFrame({
        'Código': codes,
        'Produto': products,
        'Qt Enc.': quantities
    })

    df = df.groupby(['Código', 'Produto'])[['Qt Enc.']].sum().reset_index()
    df['Código'] = df['Código'].astype(str)
    df = df.sort_values(by='Produto', ascending=True)

    return df

# ===========================================Funcoes para Unir Sell Outs s de Encomenda SIFARMA=================================


def cria_e_transforma_dataframe(ficheiro):
    df = pd.read_csv(ficheiro, encoding='latin1', sep=';')
    # Step 1: Get the value in column "localização" for the row with the biggest value in column "Tuni"
    max_tuni_loc = df.loc[df['T Uni'].idxmax(), 'LOCALIZACAO']
    # Step 2: Filter the dataframe only for rows where "localização" has the value obtained above
    df = df[df['LOCALIZACAO'] == max_tuni_loc]
    df['LOCALIZACAO'] = ficheiro.name.split('.')[0]
    df['CÓDIGO'] = df['CÓDIGO'].astype(str)

    # Formatar PVP
    df['PVP'] = df['PVP'].str.replace(',', '.')
    df['PVP'] = df['PVP'].astype(float).round(2)

    # FORMATAR P.CUSTO (NOVA LINHA)
    df['P.CUSTO'] = df['P.CUSTO'].str.replace(',', '.')
    df['P.CUSTO'] = df['P.CUSTO'].astype(float).round(2)

    df = df.drop(columns=['STOCK TOT',  'VALOR VENDAS', 'MG (%)'])

    # Mudar a posicao da Coluna PVP para depois da coluna Localização
    colunas = list(df.columns)
    coluna_a_mover_pvp = colunas.pop(colunas.index('PVP'))
    coluna_a_mover_pcusto = colunas.pop(colunas.index('P.CUSTO'))

    colunas.insert(3, coluna_a_mover_pvp)
    colunas.insert(4, coluna_a_mover_pcusto)  # Inserir P.CUSTO depois de PVP

    df = df[colunas]

    return df


def sellout_total(dataframe_combinada, df_master_produtos):

    # 0. Filtro INICIAL: Remover linhas sem stock E sem vendas antes de calcular qualquer média

    # Isto evita que preços antigos de produtos inativos distorçam a média do grupo

    filtro = (dataframe_combinada['STOCK'] != 0) | (
        dataframe_combinada['T Uni'] != 0)

    dataframe_combinada = dataframe_combinada[filtro].copy()

    # Ajustar índice para ignorar CÓDIGO, LOCALIZACAO, PVP, P.CUSTO de forma robusta

    # Em vez de slicing por indice [4:], vamos excluir explicitamente as colunas que NAO queremos somar

    colunas_nao_somar = ['CÓDIGO', 'DESIGNAÇÃO', 'LOCALIZACAO',
                         'PVP', 'P.CUSTO', 'PVP_Médio', 'P.CUSTO_Médio']

    colunas_agregar = [
        col for col in dataframe_combinada.columns if col not in colunas_nao_somar]

    # 1. Agregar Vendas (Soma) por CÓDIGO

    grouped_df = dataframe_combinada.groupby(
        'CÓDIGO')[colunas_agregar].sum().reset_index()

    # 2. Calcular média de PVP por CÓDIGO

    pvp_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['PVP'].mean().round(2).reset_index()

    # 3. Calcular média de P.CUSTO por CÓDIGO

    # (Mantemos o nome 'P.CUSTO' para o concat)

    pcusto_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['P.CUSTO'].mean().round(2).reset_index()

    # 4. Juntar tudo por CÓDIGO

    grouped_df = pd.merge(grouped_df, pvp_medio_df, on='CÓDIGO', how='left')

    grouped_df = pd.merge(grouped_df, pcusto_medio_df, on='CÓDIGO', how='left')

    # 5. Adicionar a designação limpa a partir do df_master_produtos

    grouped_df = pd.merge(grouped_df, df_master_produtos,
                          on='CÓDIGO', how='left')

    # 5.1 Garantir Maiúsculas

    grouped_df['DESIGNAÇÃO'] = grouped_df['DESIGNAÇÃO'].str.title()

    # 6. Renomear colunas de preço para a vista agrupada

    grouped_df = grouped_df.rename(
        columns={'PVP': 'PVP_Médio', 'P.CUSTO': 'P.CUSTO_Médio'})

    # 7. Reordenar colunas (DESIGNAÇÃO, PVP_Médio, P.CUSTO_Médio)

    colunas = list(grouped_df.columns)

    col_designacao = colunas.pop(colunas.index('DESIGNAÇÃO'))

    col_pvp = colunas.pop(colunas.index('PVP_Médio'))

    col_pcusto = colunas.pop(colunas.index('P.CUSTO_Médio'))

    colunas.insert(1, col_designacao)  # Depois de CÓDIGO

    colunas.insert(2, col_pvp)        # Depois de DESIGNAÇÃO

    colunas.insert(3, col_pcusto)     # Depois de PVP_Médio

    grouped_df = grouped_df[colunas]

    # 8. Filtro final e ordenação com base na DESIGNAÇÃO limpa

    grouped_df = grouped_df.sort_values(by='DESIGNAÇÃO', ascending=True)

    return grouped_df


def combina_e_agrega_df(dataframe_combinada, df_master_produtos):

    # 0. Filtro INICIAL: Remover linhas sem stock E sem vendas antes de calcular qualquer média
    # Isto evita que preços antigos de produtos inativos distorçam a média do grupo
    filtro = (dataframe_combinada['STOCK'] != 0) | (
        dataframe_combinada['T Uni'] != 0)
    dataframe_combinada = dataframe_combinada[filtro].copy()

    # Ajustar índice para ignorar CÓDIGO, LOCALIZACAO, PVP, P.CUSTO de forma robusta
    # Em vez de slicing por indice [4:], vamos excluir explicitamente as colunas que NAO queremos somar
    colunas_nao_somar = ['CÓDIGO', 'DESIGNAÇÃO', 'LOCALIZACAO',
                         'PVP', 'P.CUSTO', 'PVP_Médio', 'P.CUSTO_Médio']
    colunas_agregar = [
        col for col in dataframe_combinada.columns if col not in colunas_nao_somar]

    # 1. Agregar Vendas (Soma) por CÓDIGO
    grouped_df = dataframe_combinada.groupby(
        'CÓDIGO')[colunas_agregar].sum().reset_index()

    # 2. Calcular média de PVP por CÓDIGO
    # (Mantemos o nome 'PVP' para o concat)
    pvp_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['PVP'].mean().round(2).reset_index()

    # 3. Calcular média de P.CUSTO por CÓDIGO
    # (Mantemos o nome 'P.CUSTO' para o concat)
    pcusto_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['P.CUSTO'].mean().round(2).reset_index()

    # 4. Juntar tudo por CÓDIGO (sem a designação ainda)
    grouped_df = pd.merge(grouped_df, pvp_medio_df, on='CÓDIGO', how='left')
    grouped_df = pd.merge(grouped_df, pcusto_medio_df, on='CÓDIGO', how='left')

    # 5. Adicionar 'LOCALIZACAO' à linha de total
    grouped_df['LOCALIZACAO'] = 'Zgrupo_Total'

    # 6. Concatenar a dataframe original (sem DESIGNAÇÃO) com o agrupado
    dataframe_combinada = pd.concat(
        [dataframe_combinada, grouped_df], ignore_index=True)

    # 7. Reintroduzir a DESIGNAÇÃO limpa a toda a dataframe
    dataframe_combinada = pd.merge(
        dataframe_combinada, df_master_produtos, on='CÓDIGO', how='left')

    # 7.1 Garantir Maiúsculas (Redundância de segurança)
    dataframe_combinada['DESIGNAÇÃO'] = dataframe_combinada['DESIGNAÇÃO'].str.title(
    )

    # 7.2 Restaurar a Ordem das Colunas (CRÍTICO para os cálculos subsequentes)
    # Move DESIGNAÇÃO para a posição 1 (logo após CÓDIGO)
    cols = list(dataframe_combinada.columns)
    # Identifica onde está a DESIGNAÇÃO (geralmente no fim após o merge)
    col_designacao_name = 'DESIGNAÇÃO'
    if col_designacao_name in cols:
        cols.remove(col_designacao_name)  # Remove da posição atual
        cols.insert(1, col_designacao_name)  # Insere na posição 1
        dataframe_combinada = dataframe_combinada[cols]  # Reordena

    # 8. Ordenar pela DESIGNAÇÃO (já limpa) e LOCALIZACAO
    dataframe_combinada = dataframe_combinada.sort_values(
        by=['DESIGNAÇÃO', 'LOCALIZACAO'], ascending=[True, True])

    return dataframe_combinada


def remover_linhas_sem_vendas_e_stock(dataframe):
    zgroup_total_zeros = dataframe[
        (dataframe['LOCALIZACAO'] == 'Zgrupo_Total')
        &
        (dataframe['STOCK'] == 0) & (dataframe['T Uni'] == 0)
    ]
    cnp_a_remover = zgroup_total_zeros['CÓDIGO'].unique()
    nova_df_filtrada = dataframe[~dataframe['CÓDIGO'].isin(cnp_a_remover)]
    return nova_df_filtrada


def aplicar_destaques(linha):
    """
    Combina três tipos de destaque:
    1. Linha total (fundo preto)
    2. Linha com observação (fundo roxo claro até 'T Uni')
    3. Célula de rutura (célula vermelha na coluna 'Proposta')
    """
    estilos = [''] * len(linha)

    # 1️⃣ Destacar linha total (PRIORIDADE MÁXIMA)
    # ⚠️ CORREÇÃO: Verificar tanto 'ZGrupo_Total' quanto 'Zgrupo_Total'
    localizacao = linha.get('LOCALIZACAO', '')
    if localizacao in ['ZGrupo_Total', 'Zgrupo_Total']:
        estilos = [
            'background-color: black; font-weight: bold; color: white'] * len(linha)
        return estilos  # Retorna imediatamente, não aplica outras formatações

    # 2️⃣ Destacar linha roxa se houver observação (DATA_OBS não nula)
    if 'DATA_OBS' in linha.index and pd.notna(linha['DATA_OBS']):
        # Encontrar o índice da coluna 'T Uni'
        if 'T Uni' in linha.index:
            idx_t_uni = linha.index.get_loc('T Uni')
            # Aplicar roxo claro do início até 'T Uni' (inclusive)
            for i in range(idx_t_uni + 1):
                # Roxo claro
                estilos[i] = 'background-color: #E6D5F5; color: black'

    # 3️⃣ Destacar célula 'Proposta' se houver rutura (SOBREPÕE o roxo se necessário)
    if 'DIR' in linha.index and pd.notna(linha['DIR']):
        if 'Proposta' in linha.index:
            idx_proposta = linha.index.get_loc('Proposta')
            estilos[idx_proposta] = 'background-color: red; color: white; font-weight: bold'

    return estilos


def formatar_excel(dataframe_final):
    """
    Formata o Excel com três tipos de destaque:
    1. Linha preta (Zgrupo_Total)
    2. Linha roxa clara (DATA_OBS não nula) - até coluna T Uni
    3. Célula vermelha (DIR não nula) - coluna Proposta
    """
    # Guardar a dataframe em excel
    output = BytesIO()
    dataframe_final.to_excel(output, index=False)

    # Carregar o ficheiro Excel em memória com openpyxl
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # ========== ESTILOS ==========
    # Estilo linha total (preto)
    font_total = Font(bold=True, color="FFFFFF")
    fill_total = PatternFill(start_color="000000",
                             end_color="000000", fill_type="solid")

    # Estilo linha com observação (roxo claro)
    fill_roxo = PatternFill(start_color="E6D5F5",
                            end_color="E6D5F5", fill_type="solid")
    font_roxo = Font(color="000000")

    # Estilo célula rutura (vermelho)
    fill_vermelho = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_vermelho = Font(bold=True, color="FFFFFF")

    # ========== IDENTIFICAR COLUNAS ==========
    # Ler cabeçalhos da primeira linha
    headers = [cell.value for cell in ws[1]]

    # Encontrar índices das colunas (adicionar 1 porque Excel começa em 1)
    try:
        col_localizacao = headers.index('LOCALIZACAO') + 1
    except ValueError:
        col_localizacao = None

    try:
        col_data_obs = headers.index('DATA_OBS') + 1
    except ValueError:
        col_data_obs = None

    try:
        col_dir = headers.index('DIR') + 1
    except ValueError:
        col_dir = None

    try:
        col_proposta = headers.index('Proposta') + 1
    except ValueError:
        col_proposta = None

    try:
        col_t_uni = headers.index('T Uni') + 1
    except ValueError:
        col_t_uni = None

    # ========== APLICAR FORMATAÇÕES ==========
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):

        # 1️⃣ LINHA TOTAL (prioridade máxima)
        if col_localizacao and row[col_localizacao - 1].value in ['ZGrupo_Total', 'Zgrupo_Total']:
            for cell in row:
                cell.font = font_total
                cell.fill = fill_total
            continue  # Não aplica outras formatações

        # 2️⃣ LINHA ROXA (se DATA_OBS não for None)
        if col_data_obs and row[col_data_obs - 1].value is not None:
            # Aplicar roxo do início até T Uni (inclusive)
            if col_t_uni:
                for col_idx in range(1, col_t_uni + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_roxo
                    cell.font = font_roxo

        # 3️⃣ CÉLULA VERMELHA (se DIR não for None)
        if col_dir and col_proposta and row[col_dir - 1].value is not None:
            cell_proposta = ws.cell(row=row_idx, column=col_proposta)
            cell_proposta.fill = fill_vermelho
            cell_proposta.font = font_vermelho

    # Guardar o ficheiro de volta no buffer de memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def unir_sell_out_com_esgotados(df_sell_out, df_esgotados):
    """
    Funcao responsável por se existir dataframe de esgotados, vais adicionar as colunas de Data de Inicio e Data de fim
    Temporariamente ficará a coluna de timedelta para fazer calculo, depois será removida
    """
    df_completa = df_sell_out.merge(
        df_esgotados[['Número de registo', 'Data de início de rutura', 'Data prevista para reposição', 'TimeDelta']], left_on='CÓDIGO', right_on='Número de registo', how='left',
    )
    df_completa.drop(columns='Número de registo', axis=1, inplace=True)

    return df_completa


def calcular_proposta_esgotados(df, col_media='Media', col_timedelta='TimeDelta', col_stock='STOCK', col_proposta='Proposta'):
    """
    Calcula o valor da coluna 'Proposta' de acordo com as regras:
    - Se 'TimeDelta' for None/NaN, mantém o valor atual de 'Proposta'.
    - Caso contrário, novo valor = ('Média' / 30) * 'TimeDelta'
    """

    df[col_timedelta] = pd.to_numeric(df[col_timedelta], errors='coerce')

    mask = df[col_timedelta].notna()

    df.loc[mask, col_proposta] = (
        (df.loc[mask, col_media] / 30) *
        df.loc[mask, col_timedelta] - df.loc[mask, col_stock]
    ).round(0).astype(int)

    return df


def load_produtos_nao_comprar(url_google_sheets=GOOGLE_SHEETS):
    """
    Funcao que a partir da URL da google sheets, obtem uma dataframe
    organiza e remove duplicados por forma a ficar sempre com a informacao mais recente"""
    try:
        nc_df = pd.read_excel(GOOGLE_SHEETS)
        # Uniformizar e garantir que os nomes das farmácuas ficam sempre em title
        nc_df['FARMACIA'] = nc_df['FARMACIA'].str.title()
        # Transformar a coluna CNP numa string para poder fazer a junção com a dataframe do sell out
        nc_df['CNP'] = nc_df['CNP'].astype(str)

        # Garantir que em cado de duplicados ficamos sempre com a linha mais recente

        # Inicialmente de fforma explicitar garantir que a coluna Data está no formato datetime
        nc_df['DATA'] = pd.to_datetime(nc_df['DATA'], format='%d-%m-%Y')
        # Ordena por CNP, FARMACIA e DATA (da mais recente para a mais antiga)
        nc_df = nc_df.sort_values(
            by=['CNP', 'FARMACIA', 'DATA'], ascending=[True, True, False])
        # Remove duplicados mantendo a mais recente (graças à ordenação)
        nc_df = nc_df.drop_duplicates(
            subset=['CNP', 'FARMACIA'], keep='first').reset_index(drop=True)
        return nc_df
    except:
        st.warning(
            "⚠️ Nao foi possível carregar a lista de produtos a não comprar.")
        return pd.DataFrame(columns=['CNP', 'FARMACIA', 'DATA'])


def unir_df_na_comprar_a_df_clean(df_clean, df_nao_comprar):
    """ Funcao responsável por depois de obter a dataframe a partir do excel do google sheets
    faz o merge tendo por base o CNP e tammbem o nome da FARMACIAD
    Depois do merge feito, apenas fica com a Data da infomração porque mais nada será relevante
    """
    df_clean = df_clean.merge(df_nao_comprar, left_on=[
                              'CÓDIGO', 'LOCALIZACAO'], right_on=['CNP', 'FARMACIA'], how='left')
    # Alternar o nome da coluna para ficar perceptível
    df_clean.rename(columns={'DATA': 'DATA_OBS'}, inplace=True)
    # remover as colunas CNP e FARMACIA QUE NAO TEM INTERESSE
    df_clean.drop(columns=['CNP', 'FARMACIA'], axis=1, inplace=True)
    return df_clean


def processar_logica_negocio(df, df_esgotados, df_nao_comprar, cols_selecionadas, pesos, valor_previsao, agrupado=False):
    """
    Aplica a lógica de negócio central:
    1. Cálculo da Média Ponderada
    2. Cálculo da Proposta de Encomenda
    3. Integração com BD de Esgotados (Infarmed)
    4. Integração com Lista de Não Comprar
    """
    if df.empty:
        return df

    df = df.copy()

    # 1. Média Ponderada e Proposta Inicial
    try:
        df['Media'] = df[cols_selecionadas].dot(pesos)
        df['Proposta'] = (df['Media'] * valor_previsao -
                          df['STOCK']).round(0).astype(int)
    except Exception as e:
        st.error(
            f"Erro ao calcular médias. Verifique se o número de colunas de meses corresponde aos pesos definidos. Detalhe: {e}")
        return df

    # 2. Integração com Esgotados (Infarmed)
    if not df_esgotados.empty:
        try:
            df = unir_sell_out_com_esgotados(df, df_esgotados)
            df = calcular_proposta_esgotados(df)

            # Formatar datas
            if 'Data de início de rutura' in df.columns:
                df['Data de início de rutura'] = pd.to_datetime(
                    df['Data de início de rutura']).dt.strftime('%d-%m-%Y')
            if 'Data prevista para reposição' in df.columns:
                df['Data prevista para reposição'] = pd.to_datetime(
                    df['Data prevista para reposição']).dt.strftime('%d-%m-%Y')

            df.rename(columns={'Data de início de rutura': 'DIR',
                      'Data prevista para reposição': 'DPR'}, inplace=True)

            # Remover coluna auxiliar TimeDelta
            if 'TimeDelta' in df.columns:
                df.drop(columns='TimeDelta', axis=1, inplace=True)
        except Exception as e:
            st.warning(
                f"⚠️ Ocorreu um erro ao unir a dataframe de esgotados. {str(e)}")

    # 3. Integração com Produtos a Não Comprar
    if not df_nao_comprar.empty:
        try:
            if agrupado:
                # Lógica Agrupada: Merge apenas por CÓDIGO
                # Criar master de CNPs proibidos (remove duplicados para não multiplicar linhas)
                df_nc_unique = df_nao_comprar[['CNP', 'DATA']].sort_values(
                    'DATA', ascending=False).drop_duplicates(subset=['CNP'], keep='first')

                df = df.merge(df_nc_unique, left_on='CÓDIGO',
                              right_on='CNP', how='left')

                # Renomear e limpar
                df.rename(columns={'DATA': 'DATA_OBS'}, inplace=True)
                if 'CNP' in df.columns:
                    df.drop(columns=['CNP'], axis=1, inplace=True)
            else:
                # Lógica Detalhada: Merge por CÓDIGO e LOCALIZACAO (FARMACIA)
                df = unir_df_na_comprar_a_df_clean(df, df_nao_comprar)
        except Exception as e:
            st.warning(f"⚠️ Erro ao integrar produtos a não comprar: {str(e)}")

    # Remover coluna auxiliar da média se desejar limpar (opcional, mantendo coerência com código anterior)
    if 'Media' in df.columns:
        df.drop(columns='Media', axis=1, inplace=True)

    return df


# Title for your app
st.title("Sistema de Agrupamento de SellOuts")

# Botao para limpar a chahe e reiniciar o programa
if st.button("🧹 Limpar dados carregados"):
    # Recria DataFrames vazias
    st.session_state["dataframe_total"] = pd.DataFrame()
    st.session_state["dataframe_final"] = pd.DataFrame()
    st.session_state["dataframe_clean"] = pd.DataFrame()
    st.session_state["dataframe_agrupada"] = pd.DataFrame()

    # 🔸 Limpa ficheiros de upload (widgets)
    # Incrementa a chave para forçar o uploader a resetar
    st.session_state['uploader_key'] += 1
    # st.session_state.pop("uploader_csv", None) # Esta linha já não é necessária

    # Opcional: limpa cache temporário (sem tocar nos decoradores de BD)
    st.cache_data.clear()

    # Força recarregamento da app (reinicia a execução)
    st.rerun()


# Funcao que vai tentar obter a base de dados de esgotados do Pharmaruptures
df_esgotados, data_consulta = obter_base_dados_esgotados()

# Adcionar um cartao com a data da Consulta da BD
st.markdown(
    f"""
    <div style="
        display:flex;
        align-items:center;
        justify-content:center;
        background:linear-gradient(135deg, #e0f7fa, #f1f8e9);
        padding:15px;
        border-radius:15px;
        box-shadow:0 2px 10px rgba(0,0,0,0.1);
    ">
        <span style="font-size:24px;margin-right:10px;">🗓️</span>
        <div>
            <div style="font-size:16px;color:#555;">Data Consulta BD Rupturas</div>
            <div style="font-size:24px;font-weight:bold;color:#0078D7;">{data_consulta}</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

# Obter os produtos a nao comprar
df_nao_comprar = load_produtos_nao_comprar()


# Calcular o verdadeiro time delta entre o dia corrente e a data prevista de reposição
if not df_esgotados.empty:
    # Garantir que a coluna data prevista +e uma coluna datetime
    df_esgotados['Data prevista para reposição'] = pd.to_datetime(
        df_esgotados['Data prevista para reposição'])
    # Calcula TimeDelta = data de fim - hoje
    hoje = pd.Timestamp(datetime.now().date())

    # Recalcula o Time Delta entre a data de hoje e a data prevista de reposicao em dias
    df_esgotados['TimeDelta'] = (
        df_esgotados['Data prevista para reposição'] - hoje).dt.days


# Upload Ficheiros Cria e agrupa numa só dataframe
uploaded_file = st.file_uploader("Choose a file",  accept_multiple_files=True,
                                 type='csv', key=f"uploader_csv_{st.session_state['uploader_key']}")
if uploaded_file:
    lista_df = []
    for ficheiro in uploaded_file:
        dataframe_temp = cria_e_transforma_dataframe(ficheiro)
        lista_df.append(dataframe_temp)

    dataframe_total = pd.concat(lista_df, ignore_index=True)

    # 1. Criar Master de Produtos com CÓDIGO e DESIGNAÇÃO
    df_master_produtos = dataframe_total[['CÓDIGO', 'DESIGNAÇÃO']].copy()

    # 2. Limpar DESIGNAÇÃO no master (remover acentos e caracteres especiais, converter para maiúsculas)
    df_master_produtos['DESIGNAÇÃO'] = df_master_produtos['DESIGNAÇÃO'].apply(
        limpar_designacao)

    # 3. Remover duplicados por CÓDIGO, mantendo a primeira DESIGNAÇÃO limpa
    df_master_produtos = df_master_produtos.drop_duplicates(
        subset='CÓDIGO', keep='first').reset_index(drop=True)

    # 4. Remover a coluna DESIGNAÇÃO da dataframe_total antes de processar
    dataframe_total = dataframe_total.drop(columns=['DESIGNAÇÃO'])

    # .copy() para evitar SettingWithCopyWarning
    dataframe_final = combina_e_agrega_df(
        dataframe_total.copy(), df_master_produtos)

    dataframe_clean = remover_linhas_sem_vendas_e_stock(dataframe_final)
    dataframe_clean = dataframe_clean.rename(columns={'PVP': 'PVP_Médio'})

    # Para garantir que os nomes das localizações ficarão sempre uniforme em title
    dataframe_clean['LOCALIZACAO'] = dataframe_clean['LOCALIZACAO'].str.title()

    # .copy() para evitar SettingWithCopyWarning
    dataframe_agrupada = sellout_total(
        dataframe_total.copy(), df_master_produtos)

    # Aplicar estilo

    # dataframe_destaque = dataframe_clean.style.apply(aplicar_destaques, axis=1).format({'PVP_Médio': '{:.2f}'})

    #  Toggle para saber a partir de que mes se vai calcular a média ponderada -> Mes acutal ou anterior
    anterior = st.toggle("Média Ponderada com Base no mês ANTERIOR?")

    # Pesos e indices relativos para calcular e adicionar a coluna de média ponderada
    if anterior:
        pesos = [0.4, 0.3, 0.2, 0.1]
        indice_colunas = [-3, -4, -5, -6]
    else:
        pesos = [0.4, 0.3, 0.2, 0.1]
        indice_colunas = [-2, -3, -4, -5]

    # Pega os nomes das colunas dinamicamente
    cols_selecionadas = [dataframe_clean.columns[i] for i in indice_colunas]

    # Toggle para ver detalhes ou agrupada
    on = st.toggle("Ver Detalhe de Sell Out?")

    # Número mínimo 1, máximo 4, passo 0.1 para permitir decimais
    st.subheader(
        "Indicar o numero de meses a prever a encomenda (pode ser número décimal. Ex. 1.5 meses)")
    valor = st.number_input(label="Meses a prever",  # Label para acessibilidade
                            label_visibility="collapsed",  # Esconde o label visualmente
                            min_value=1.0,
                            max_value=4.0,
                            value=1.0,      # valor inicial
                            step=0.1,       # passo do incremento/decremento
                            format="%.1f"   # mostra sempre 1 decimal
                            )

    st.write(f"A Preprarar encomenda para {valor:.1f} Meses")

    if on:
        # === VISTA DETALHADA ===
        dataframe_clean = processar_logica_negocio(
            dataframe_clean,
            df_esgotados,
            df_nao_comprar,
            cols_selecionadas,
            pesos,
            valor,
            agrupado=False
        )

        # Aplicar estilo visual no Streamlit
        # Formatamos PVP e P.Custo para 2 casas decimais na visualização
        dataframe_destaque = dataframe_clean.style.apply(
            aplicar_destaques, axis=1).format({'PVP_Médio': '{:.2f}', 'P.CUSTO': '{:.2f}'})
        st.dataframe(dataframe_destaque, hide_index=True, width='stretch')

        excel_formatado = formatar_excel(dataframe_clean)

    else:
        # === VISTA AGRUPADA ===
        dataframe_agrupada = processar_logica_negocio(
            dataframe_agrupada,
            df_esgotados,
            df_nao_comprar,
            cols_selecionadas,
            pesos,
            valor,
            agrupado=True
        )

        # Aplicar estilo visual no Streamlit
        # Formatamos PVP e P.Custo para 2 casas decimais na visualização
        dataframe_destaque_agrupado = dataframe_agrupada.style.apply(
            aplicar_destaques, axis=1).format({'PVP_Médio': '{:.2f}', 'P.CUSTO_Médio': '{:.2f}'})
        st.dataframe(dataframe_destaque_agrupado,
                     hide_index=True, width='stretch')

        excel_formatado = formatar_excel(dataframe_agrupada)

    # excel_formatado = formatar_excel(dataframe_clean)

    # Botão para download no Streamlit
    st.download_button(
        label="Download Excel",
        data=excel_formatado,
        file_name='Sell_Out_GRUPO.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# Toggle para ver detalhes ou agrupada
on_unir_encomendas = st.toggle("Activar funcao de Unir encomendas?")

if on_unir_encomendas:
    # Title for your app
    st.subheader('Unir Notas de Encomenda Sifarma - > Formatos em PDF')

    # Upload Ficheiros Cria e agrupa numa só dataframe
    pdf_notas_encomenda = st.file_uploader(
        "Selecionar as Notas de Encomenda Sifarma (PDF)",  accept_multiple_files=True, type='pdf')

    if pdf_notas_encomenda:
        texto_total = ''
        for encomenda in pdf_notas_encomenda:
            # print(type(encomenda))
            texto_total += extract_text_from_pdf(encomenda)

        encomenda_final_df = extract_data_from_text(texto_total)

        st.dataframe(encomenda_final_df, hide_index=True,)


# st.divider()

# st.text('DATAFRAME ESGOTADOS')


# st.write(df_esgotados)


# st.divider()
# st.text('DATAFRAME COM ESGOTADDOS')


# st.divider()
# st.text('PRODUTOS A NAO COMRPAR')
# st.write(df_nao_comprar)
