import streamlit as st
import pandas as pd
import json
import os
import unicodedata
from dotenv import load_dotenv
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from processar_infoprex import ler_ficheiro_infoprex, extrair_codigos_txt
from stockreorder import gerar_plano_redistribuicao

# Carregar variáveis de ambiente
load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")
GOOGLE_SHEETS = os.getenv("GOOGLE_SHEETS")

# Configuração da Página - Deve ser a primeira instrução
st.set_page_config(
    page_title="Orders Master Infoprex",
    page_icon="📦",
    layout='wide',
    initial_sidebar_state="expanded"
)

pd.options.display.float_format = '{:.2f}'.format

# ==========================================
# Funções Auxiliares de Mapeamento
# ==========================================


@st.cache_data
def carregar_localizacoes():
    """Carrega o dicionário de localizações do ficheiro JSON."""
    try:
        if os.path.exists('localizacoes.json'):
            with open('localizacoes.json', 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        st.sidebar.error(f"Erro ao carregar localizacoes.json: {e}")
        return {}


def mapear_localizacao(nome, dict_locs):
    """
    Procura chaves do dicionário (em minúsculas) dentro do nome original (em minúsculas).
    Se encontrar, devolve o valor associado em Title Case.
    Se não encontrar, devolve o nome original em Title Case.
    """
    if not isinstance(nome, str):
        return nome
    nome_lower = nome.lower()
    for chave, valor in dict_locs.items():
        if chave.lower() in nome_lower:
            return valor.title()
    return nome.title()

# ==========================================
# Funções de Dados Externos (Caches)
# ==========================================


@st.cache_data(show_spinner="🔄 A carregar dados da base de dados de Esgotados...", ttl=3600)
def obter_base_dados_esgotados():
    """Função que obtém a base de dados de Esgotados a partir de um URL (Excel/Google Sheets)."""
    df_esgotados = pd.DataFrame()
    data_consulta = 'Não foi possível carregar a INFO'

    if not DATABASE_URL:
        st.sidebar.warning("⚠️ DATABASE_URL não definida no .env")
        return df_esgotados, data_consulta

    try:
        df_esgotados = pd.read_excel(DATABASE_URL)
        colunas = ['Número de registo', 'Nome do medicamento', 'Data de início de rutura',
                   'Data prevista para reposição', 'TimeDelta', 'Data da Consulta']
        df_esgotados = df_esgotados[colunas].copy()

        df_esgotados['Número de registo'] = df_esgotados['Número de registo'].astype(
            str)
        df_esgotados['TimeDelta'] = pd.to_numeric(
            df_esgotados['TimeDelta'], errors='coerce')

        data_consulta = str(df_esgotados["Data da Consulta"].iloc[0])[:10]

        # Calcular o verdadeiro time delta entre o dia corrente e a data prevista de reposição
        df_esgotados['Data prevista para reposição'] = pd.to_datetime(
            df_esgotados['Data prevista para reposição'])
        hoje = pd.Timestamp(datetime.now().date())
        df_esgotados['TimeDelta'] = (
            df_esgotados['Data prevista para reposição'] - hoje).dt.days

        return df_esgotados, data_consulta
    except Exception as e:
        print(f"Erro BD Esgotados: {e}")
        return pd.DataFrame(), data_consulta


@st.cache_data(show_spinner="🔄 A carregar lista de Produtos a Não Comprar...", ttl=3600)
def load_produtos_nao_comprar(_dict_locs):
    """Função que obtém a lista de produtos a não comprar do Google Sheets."""
    if not GOOGLE_SHEETS:
        return pd.DataFrame(columns=['CNP', 'FARMACIA', 'DATA'])

    try:
        nc_df = pd.read_excel(GOOGLE_SHEETS)
        nc_df['FARMACIA'] = nc_df['FARMACIA'].apply(
            lambda x: mapear_localizacao(x, _dict_locs))
        nc_df['CNP'] = nc_df['CNP'].astype(str)
        nc_df['DATA'] = pd.to_datetime(nc_df['DATA'], format='%d-%m-%Y')
        nc_df = nc_df.sort_values(
            by=['CNP', 'FARMACIA', 'DATA'], ascending=[True, True, False])
        nc_df = nc_df.drop_duplicates(
            subset=['CNP', 'FARMACIA'], keep='first').reset_index(drop=True)
        return nc_df
    except Exception as e:
        print(f"Erro BD Não Comprar: {e}")
        return pd.DataFrame(columns=['CNP', 'FARMACIA', 'DATA'])

# ==========================================
# Funções de Transformação e UI
# ==========================================


def limpar_designacao(texto):
    """Remove acentos, asteriscos e converte para title case para ordenação correta e consistência."""
    if not isinstance(texto, str):
        return str(texto)
    # Normaliza unicode (NFD decompõe caracteres) e remove marcas de acentuação (Mn)
    texto_limpo = ''.join(c for c in unicodedata.normalize('NFD', texto)
                          if unicodedata.category(c) != 'Mn')
    # Remove asteriscos
    texto_limpo = texto_limpo.replace('*', '')
    return texto_limpo.strip().title()


def criar_tabela_dimensao(df):
    """
    Cria a tabela mestre (dimensão) de produtos baseada na DataFrame consolidada.
    Já garantimos que os códigos começados por 1 foram removidos no processamento.
    Agrupa apenas por CÓDIGO.
    """
    if df.empty:
        return pd.DataFrame()

    df_univ_bruta = df[['CÓDIGO', 'DESIGNAÇÃO']].copy()
    df_univ_bruta['DESIGNAÇÃO'] = df_univ_bruta['DESIGNAÇÃO'].apply(
        limpar_designacao)
    df_univ = df_univ_bruta.drop_duplicates(
        subset=['CÓDIGO'], keep='first').reset_index(drop=True)

    return df_univ


def get_file_modified_time(path):
    """Obtem a data de modificacao de um ficgeiro para que se ele for modificado no 
    github ou sistema imediatamente a chache seja limpa o seja recarregado
    """
    return os.path.getmtime(path) if os.path.exists(path) else 0


@st.cache_data
def carregar_laboratorios(mtime):
    """O parametro mtime é usado apenas para invalidar a cache quando o ficheiro é modificado,
    garantindo que as atualizações são refletidas imediatamente."""

    """Carrega o dicionário de laboratórios do ficheiro JSON."""
    try:
        if os.path.exists('laboratorios.json'):
            with open('laboratorios.json', 'r', encoding='utf-8') as f:
                labs = json.load(f)
            return labs
        return {}
    except Exception as e:
        st.sidebar.error(f"Erro ao carregar laboratorios.json: {e}")
        return {}


@st.cache_data(show_spinner=False)
def processar_ficheiros_upload(ficheiros, labs_selecionados, codigos_txt, _dicionario_labs, _dict_locs):
    """
    Processa todos os ficheiros carregados, aplicando os filtros selecionados,
    elimina códigos locais (iniciados por 1), e retorna uma única dataframe.
    """
    if not ficheiros:
        return pd.DataFrame(), [], []

    # Preparar a lista de CLA baseada na seleção
    lista_cla = []
    if labs_selecionados and _dicionario_labs:
        for lab in labs_selecionados:
            lista_cla.extend(_dicionario_labs.get(lab, []))

    # Preparar a lista de Códigos
    lista_codigos = []
    if codigos_txt is not None:
        lista_codigos = extrair_codigos_txt(codigos_txt)

    lista_df = []
    erros_ficheiros = []
    for ficheiro in ficheiros:
        try:
            df_temp = ler_ficheiro_infoprex(
                ficheiro, lista_cla=lista_cla, lista_codigos=lista_codigos)
            if not df_temp.empty:
                lista_df.append(df_temp)
        except ValueError as ve:
            erros_ficheiros.append(f"Erro no ficheiro '{ficheiro.name}': {ve}")
        except Exception as e:
            erros_ficheiros.append(
                f"Erro inesperado ao processar '{ficheiro.name}': {e}")

    codigos_invalidos = []
    if lista_df:
        df_final = pd.concat(lista_df, ignore_index=True)

        # Mapear Localizações imediatamente
        df_final['LOCALIZACAO'] = df_final['LOCALIZACAO'].apply(
            lambda x: mapear_localizacao(x, _dict_locs))

        # Eliminar imediatamente códigos que começam por '1' (Códigos Locais)
        mask_local = df_final['CÓDIGO'].astype(
            str).str.strip().str.startswith('1')
        df_final = df_final[~mask_local].copy()

        if df_final.empty:
            return pd.DataFrame(), codigos_invalidos, erros_ficheiros

        # Tentar converter o CÓDIGO para numérico. O que não for número vira NaN.
        df_final['CÓDIGO_NUM'] = pd.to_numeric(
            df_final['CÓDIGO'], errors='coerce')

        # Identificar as linhas inválidas
        mask_invalid = df_final['CÓDIGO_NUM'].isna()
        if mask_invalid.any():
            codigos_invalidos = df_final.loc[mask_invalid, 'CÓDIGO'].unique(
            ).tolist()

        # Limpar a dataframe e aplicar o tipo inteiro final
        df_final = df_final[~mask_invalid].copy()
        df_final['CÓDIGO'] = df_final['CÓDIGO_NUM'].astype(int)
        df_final.drop(columns=['CÓDIGO_NUM'], inplace=True)

        return df_final, codigos_invalidos, erros_ficheiros

    return pd.DataFrame(), [], erros_ficheiros

# ==========================================
# Motores de Agregação (Vista Agrupada e Detalhada)
# ==========================================


def sellout_total(dataframe_combinada, df_master_produtos):
    # 0. Filtro INICIAL: Remover linhas sem stock E sem vendas antes de calcular qualquer média
    filtro = (dataframe_combinada['STOCK'] != 0) | (
        dataframe_combinada['T Uni'] != 0)
    dataframe_combinada = dataframe_combinada[filtro].copy()

    # Ajustar índice para ignorar CÓDIGO, DESIGNAÇÃO, LOCALIZACAO, PVP, P.CUSTO, DUC, DTVAL, CLA
    colunas_nao_somar = ['CÓDIGO', 'DESIGNAÇÃO',
                         'LOCALIZACAO', 'PVP', 'P.CUSTO', 'DUC', 'DTVAL', 'CLA']
    colunas_agregar = [
        col for col in dataframe_combinada.columns if col not in colunas_nao_somar]

    # 1. Agregar Vendas (Soma) por CÓDIGO
    grouped_df = dataframe_combinada.groupby(
        'CÓDIGO')[colunas_agregar].sum().reset_index()

    # 2. Calcular média de PVP por CÓDIGO
    pvp_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['PVP'].mean().round(2).reset_index()

    # 3. Calcular média de P.CUSTO por CÓDIGO
    pcusto_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['P.CUSTO'].mean().round(2).reset_index()

    # 4. Juntar tudo por CÓDIGO
    grouped_df = pd.merge(grouped_df, pvp_medio_df, on='CÓDIGO', how='left')
    grouped_df = pd.merge(grouped_df, pcusto_medio_df, on='CÓDIGO', how='left')

    # 5. Adicionar a designação limpa a partir do df_master_produtos
    grouped_df = pd.merge(grouped_df, df_master_produtos,
                          on='CÓDIGO', how='left')
    grouped_df['DESIGNAÇÃO'] = grouped_df['DESIGNAÇÃO'].str.title()

    # 6. Renomear colunas de preço para a vista agrupada
    grouped_df = grouped_df.rename(
        columns={'PVP': 'PVP_Médio', 'P.CUSTO': 'P.CUSTO_Médio'})

    # 7. Reordenar colunas (DESIGNAÇÃO, PVP_Médio, P.CUSTO_Médio)
    colunas = list(grouped_df.columns)
    col_designacao = colunas.pop(colunas.index('DESIGNAÇÃO'))
    col_pvp = colunas.pop(colunas.index('PVP_Médio'))
    col_pcusto = colunas.pop(colunas.index('P.CUSTO_Médio'))

    colunas.insert(1, col_designacao)
    colunas.insert(2, col_pvp)
    colunas.insert(3, col_pcusto)
    grouped_df = grouped_df[colunas]

    # 8. Filtro final e ordenação com base na DESIGNAÇÃO limpa
    grouped_df = grouped_df.sort_values(
        by=['DESIGNAÇÃO', 'CÓDIGO'], ascending=[True, True])

    return grouped_df


def combina_e_agrega_df(dataframe_combinada, df_master_produtos):
    # 0. Filtro INICIAL: Remover linhas sem stock E sem vendas
    filtro = (dataframe_combinada['STOCK'] != 0) | (
        dataframe_combinada['T Uni'] != 0)
    dataframe_combinada = dataframe_combinada[filtro].copy()

    colunas_nao_somar = ['CÓDIGO', 'DESIGNAÇÃO',
                         'LOCALIZACAO', 'PVP', 'P.CUSTO', 'DUC', 'DTVAL', 'CLA']
    colunas_agregar = [
        col for col in dataframe_combinada.columns if col not in colunas_nao_somar]

    # 1. Agregar Vendas (Soma) por CÓDIGO
    grouped_df = dataframe_combinada.groupby(
        'CÓDIGO')[colunas_agregar].sum().reset_index()

    # 2. Calcular médias
    pvp_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['PVP'].mean().round(2).reset_index()
    pcusto_medio_df = dataframe_combinada.groupby(
        'CÓDIGO')['P.CUSTO'].mean().round(2).reset_index()

    # 3. Juntar tudo
    grouped_df = pd.merge(grouped_df, pvp_medio_df, on='CÓDIGO', how='left')
    grouped_df = pd.merge(grouped_df, pcusto_medio_df, on='CÓDIGO', how='left')

    # 4. Adicionar 'LOCALIZACAO' à linha de total
    grouped_df['LOCALIZACAO'] = 'Zgrupo_Total'

    # 5. Concatenar a dataframe original (sem DESIGNAÇÃO para evitar conflitos) com o agrupado
    if 'DESIGNAÇÃO' in dataframe_combinada.columns:
        dataframe_combinada = dataframe_combinada.drop(columns=['DESIGNAÇÃO'])

    dataframe_combinada = pd.concat(
        [dataframe_combinada, grouped_df], ignore_index=True)

    # 6. Reintroduzir a DESIGNAÇÃO limpa
    dataframe_combinada = pd.merge(
        dataframe_combinada, df_master_produtos, on='CÓDIGO', how='left')
    dataframe_combinada['DESIGNAÇÃO'] = dataframe_combinada['DESIGNAÇÃO'].str.title(
    )

    # 7. Restaurar a Ordem das Colunas
    cols = list(dataframe_combinada.columns)
    col_designacao_name = 'DESIGNAÇÃO'
    if col_designacao_name in cols:
        cols.remove(col_designacao_name)
        cols.insert(1, col_designacao_name)
        dataframe_combinada = dataframe_combinada[cols]

    # Ordenação estrita por Designação, Código e Localização
    return dataframe_combinada.sort_values(by=['DESIGNAÇÃO', 'CÓDIGO', 'LOCALIZACAO'], ascending=[True, True, True])


def remover_linhas_sem_vendas_e_stock(dataframe):
    zgroup_total_zeros = dataframe[
        (dataframe['LOCALIZACAO'] == 'Zgrupo_Total') &
        (dataframe['STOCK'] == 0) & (dataframe['T Uni'] == 0)
    ]
    cnp_a_remover = zgroup_total_zeros['CÓDIGO'].unique()
    nova_df_filtrada = dataframe[~dataframe['CÓDIGO'].isin(cnp_a_remover)]
    return nova_df_filtrada

# ==========================================
# Lógica de Negócio (Propostas e Merges)
# ==========================================


def unir_sell_out_com_esgotados(df_sell_out, df_esgotados):
    df_completa = df_sell_out.merge(
        df_esgotados[['Número de registo', 'Data de início de rutura',
                      'Data prevista para reposição', 'TimeDelta']],
        left_on='CÓDIGO', right_on='Número de registo', how='left'
    )
    df_completa.drop(columns='Número de registo', axis=1,
                     inplace=True, errors='ignore')
    return df_completa


def calcular_proposta_esgotados(df, col_media='Media', col_timedelta='TimeDelta', col_stock='STOCK', col_proposta='Proposta'):
    if col_timedelta not in df.columns:
        return df
    df[col_timedelta] = pd.to_numeric(df[col_timedelta], errors='coerce')
    mask = df[col_timedelta].notna()
    df.loc[mask, col_proposta] = ((df.loc[mask, col_media] / 30) * df.loc[mask,
                                  col_timedelta] - df.loc[mask, col_stock]).round(0).astype(int)
    return df


def unir_df_na_comprar_a_df_clean(df_clean, df_nao_comprar):
    df_clean['CÓDIGO_STR'] = df_clean['CÓDIGO'].astype(str)
    df_clean = df_clean.merge(df_nao_comprar, left_on=[
                              'CÓDIGO_STR', 'LOCALIZACAO'], right_on=['CNP', 'FARMACIA'], how='left')
    df_clean.rename(columns={'DATA': 'DATA_OBS'}, inplace=True)
    df_clean.drop(columns=['CNP', 'FARMACIA', 'CÓDIGO_STR'],
                  axis=1, inplace=True, errors='ignore')
    return df_clean


def processar_logica_negocio(df, df_esgotados, df_nao_comprar, cols_selecionadas, pesos, valor_previsao, agrupado=False):
    if df.empty:
        return df
    df = df.copy()

    # 1. Média Ponderada e Proposta Inicial
    try:
        df['Media'] = df[cols_selecionadas].dot(pesos)
        df['Proposta'] = (df['Media'] * valor_previsao -
                          df['STOCK']).round(0).astype(int)
    except Exception as e:
        st.error(f"Erro ao calcular médias: {e}")
        return df

    # 2. Integração com Esgotados (Infarmed)
    if not df_esgotados.empty:
        try:
            df['CÓDIGO_STR'] = df['CÓDIGO'].astype(str)
            df_esgotados_copy = df_esgotados.copy()
            df_esgotados_copy['Número de registo'] = df_esgotados_copy['Número de registo'].astype(
                str)

            df = df.merge(
                df_esgotados_copy[['Número de registo', 'Data de início de rutura',
                                   'Data prevista para reposição', 'TimeDelta']],
                left_on='CÓDIGO_STR', right_on='Número de registo', how='left'
            )
            df.drop(columns=['Número de registo', 'CÓDIGO_STR'],
                    axis=1, inplace=True, errors='ignore')

            df = calcular_proposta_esgotados(df)

            if 'Data de início de rutura' in df.columns:
                df['Data de início de rutura'] = pd.to_datetime(
                    df['Data de início de rutura']).dt.strftime('%d-%m-%Y')
            if 'Data prevista para reposição' in df.columns:
                df['Data prevista para reposição'] = pd.to_datetime(
                    df['Data prevista para reposição']).dt.strftime('%d-%m-%Y')

            df.rename(columns={'Data de início de rutura': 'DIR',
                      'Data prevista para reposição': 'DPR'}, inplace=True)
            if 'TimeDelta' in df.columns:
                df.drop(columns='TimeDelta', axis=1, inplace=True)
        except Exception as e:
            st.warning(
                f"⚠️ Ocorreu um erro ao unir a dataframe de esgotados. {str(e)}")

    # 3. Integração com Produtos a Não Comprar
    if not df_nao_comprar.empty:
        try:
            df['CÓDIGO_STR'] = df['CÓDIGO'].astype(str)
            if agrupado:
                df_nc_unique = df_nao_comprar[['CNP', 'DATA']].sort_values(
                    'DATA', ascending=False).drop_duplicates(subset=['CNP'], keep='first')
                df = df.merge(df_nc_unique, left_on='CÓDIGO_STR',
                              right_on='CNP', how='left')
                df.rename(columns={'DATA': 'DATA_OBS'}, inplace=True)
                if 'CNP' in df.columns:
                    df.drop(columns=['CNP'], axis=1, inplace=True)
            else:
                df = df.merge(df_nao_comprar, left_on=['CÓDIGO_STR', 'LOCALIZACAO'], right_on=[
                              'CNP', 'FARMACIA'], how='left')
                df.rename(columns={'DATA': 'DATA_OBS'}, inplace=True)
                df.drop(columns=['CNP', 'FARMACIA'], axis=1,
                        inplace=True, errors='ignore')
            df.drop(columns=['CÓDIGO_STR'], axis=1,
                    inplace=True, errors='ignore')
        except Exception as e:
            st.warning(f"⚠️ Erro ao integrar produtos a não comprar: {str(e)}")

    if 'Media' in df.columns:
        df.drop(columns='Media', axis=1, inplace=True)

    return df


def aplicar_destaques(linha):
    estilos = [''] * len(linha)
    localizacao = linha.get('LOCALIZACAO', '')
    if localizacao in ['ZGrupo_Total', 'Zgrupo_Total']:
        return ['background-color: black; font-weight: bold; color: white'] * len(linha)

    if 'DATA_OBS' in linha.index and pd.notna(linha['DATA_OBS']):
        if 'T Uni' in linha.index:
            idx_t_uni = linha.index.get_loc('T Uni')
            for i in range(idx_t_uni + 1):
                estilos[i] = 'background-color: #E6D5F5; color: black'

    if 'DIR' in linha.index and pd.notna(linha['DIR']):
        if 'Proposta' in linha.index:
            idx_proposta = linha.index.get_loc('Proposta')
            estilos[idx_proposta] = 'background-color: red; color: white; font-weight: bold'

    return estilos


def formatar_excel(dataframe_final):
    output = BytesIO()
    dataframe_final.to_excel(output, index=False)
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    font_total = Font(bold=True, color="FFFFFF")
    fill_total = PatternFill(start_color="000000",
                             end_color="000000", fill_type="solid")
    fill_roxo = PatternFill(start_color="E6D5F5",
                            end_color="E6D5F5", fill_type="solid")
    font_roxo = Font(color="000000")
    fill_vermelho = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_vermelho = Font(bold=True, color="FFFFFF")

    headers = [cell.value for cell in ws[1]]
    col_localizacao = headers.index(
        'LOCALIZACAO') + 1 if 'LOCALIZACAO' in headers else None
    col_data_obs = headers.index('DATA_OBS') + \
        1 if 'DATA_OBS' in headers else None
    col_dir = headers.index('DIR') + 1 if 'DIR' in headers else None
    col_proposta = headers.index('Proposta') + \
        1 if 'Proposta' in headers else None
    col_t_uni = headers.index('T Uni') + 1 if 'T Uni' in headers else None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        if col_localizacao and row[col_localizacao - 1].value in ['ZGrupo_Total', 'Zgrupo_Total']:
            for cell in row:
                cell.font = font_total
                cell.fill = fill_total
            continue
        if col_data_obs and row[col_data_obs - 1].value is not None:
            if col_t_uni:
                for col_idx in range(1, col_t_uni + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_roxo
                    cell.font = font_roxo
        if col_dir and col_proposta and row[col_dir - 1].value is not None:
            cell_proposta = ws.cell(row=row_idx, column=col_proposta)
            cell_proposta.fill = fill_vermelho
            cell_proposta.font = font_vermelho

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# Sidebar e Main App
# ==========================================


def render_sidebar():
    st.sidebar.header("⚙️ Configurações e Filtros")

    st.sidebar.subheader("1. Filtrar por Laboratório")
    st.sidebar.markdown(
        "<small>Selecione um ou mais laboratórios para filtrar (Ignorado se usar txt abaixo).</small>", unsafe_allow_html=True)

    dicionario_labs = carregar_laboratorios(
        get_file_modified_time('laboratorios.json'))
    lista_nomes_labs = sorted(list(dicionario_labs.keys()))

    labs_selecionados = st.sidebar.multiselect(
        "Laboratórios disponíveis:",
        options=lista_nomes_labs,
        default=None,
        placeholder="Escolha os laboratórios..."
    )

    st.sidebar.subheader("2. Filtrar por Códigos (Prioridade)")
    st.sidebar.markdown(
        "<small>Carregue um ficheiro .txt com um código por linha. Se usar isto, o filtro por laboratório será ignorado.</small>", unsafe_allow_html=True)
    ficheiro_codigos = st.sidebar.file_uploader("Ficheiro de Códigos (.txt)", type=[
                                                'txt'], key="upload_txt_codigos")

    st.sidebar.divider()

    st.sidebar.subheader("3. Dados Base (Infoprex)")
    ficheiros_infoprex = st.sidebar.file_uploader(
        "Ficheiros de Vendas Sifarma (.txt)",
        accept_multiple_files=True,
        type=['txt'],
        key="uploader_infoprex"
    )

    processar = st.sidebar.button(
        "🚀 Processar Dados", width='stretch', type="primary")

    return {
        'ficheiro_codigos': ficheiro_codigos,
        'labs_selecionados': labs_selecionados,
        'dicionario_labs': dicionario_labs,
        'ficheiros_infoprex': ficheiros_infoprex,
        'processar': processar
    }


def render_documentacao():
    st.markdown("""
    Este sistema foi desenhado para agrupar, analisar e sugerir propostas de encomendas baseadas em ficheiros exportados do **Sifarma (Novo Módulo Infoprex)**.
    
    ### 📂 1. Formatos de Ficheiros Suportados
    *   **Ficheiros Vendas Infoprex (`.txt`):** Ficheiros extraídos do novo módulo Sifarma. Devem incluir as colunas necessárias (CPR, NOM, Vendas, etc.). Estes ficheiros contêm as vendas de uma farmácia específica. Pode carregar múltiplos ficheiros em simultâneo. Se inserir o ficheiro errado aqui, o sistema emitirá um alerta amigável.
    *   **Ficheiro de Códigos Filtro (`.txt`):** Um ficheiro de texto simples, idealmente com **um código numérico (CNP/CPR) por linha**. O sistema ignora automaticamente cabeçalhos de texto e linhas em branco.
    
    ### ⚙️ 2. Como Filtrar os Dados (Sidebar)
    A filtragem é aplicada **antes** de os ficheiros serem processados, o que poupa tempo e memória. Além disso, o sistema garante que **todos os códigos na tabela final são inteiros**. Quaisquer códigos impossíveis de converter (ex: produtos não registados) serão eliminados com um aviso de segurança. *Códigos de farmácia locais (iniciados por 1) são automaticamente descartados.*
    *   **Prioridade Máxima:** Se carregar o **Ficheiro de Códigos (.txt)**, o sistema ignora qualquer outra seleção e foca-se apenas nesses produtos.
    *   **Seleção de Laboratórios:** Caso não use o ficheiro TXT, pode selecionar um ou múltiplos laboratórios no menu dropdown. O sistema usará os códigos internos da coluna `CLA` associados a cada laboratório para filtrar a lista.
    
    ### 🔄 3. Workflow do Processo
    1.  Abra a **Sidebar** (menu lateral à esquerda).
    2.  (Opcional) Selecione os Laboratórios pretendidos.
    3.  (Opcional) Carregue o ficheiro `.txt` com os CNPs específicos a encomendar.
    4.  Carregue os múltiplos ficheiros `.txt` do Infoprex (um por farmácia).
    5.  Ajuste as opções de negócio na página principal (Média Ponderada, Meses a prever).
    6.  Clique no botão azul **🚀 Processar Dados**.
    7.  Veja a tabela consolidada gerada abaixo.
    
    ---
    *Precisa de adicionar mais laboratórios ou mapeamentos de farmácias?*  
    Edite os ficheiros `laboratorios.json` e `localizacoes.json` na pasta do sistema para atualizar as listas automaticamente.
    """)

# ==========================================
# Estrutura Principal da App
# ==========================================


def main():
    st.title("📦 Orders Master Infoprex")

    # 0. Carregar mapeamento de localizações
    dict_locs = carregar_localizacoes()

    # Carregar BDs Externas
    df_esgotados, data_consulta = obter_base_dados_esgotados()
    df_nao_comprar = load_produtos_nao_comprar(dict_locs)

    # Informação Data da BD Rupturas
    st.markdown(
        f"""
        <div style="
            display:flex;
            align-items:center;
            justify-content:center;
            background:linear-gradient(135deg, #e0f7fa, #f1f8e9);
            padding:15px;
            border-radius:15px;
            box-shadow:0 2px 10px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        ">
            <span style="font-size:24px;margin-right:10px;">🗓️</span>
            <div>
                <div style="font-size:16px;color:#555;">Data Consulta BD Rupturas</div>
                <div style="font-size:24px;font-weight:bold;color:#0078D7;">{data_consulta}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # Documentação Expansível
    with st.expander("ℹ️ Precisa de ajuda? Clique aqui para ler a Documentação e Workflow do Sistema"):
        render_documentacao()

    # Inicializa o estado dos últimos filtros aplicados (necessário inicializar antes de usar)
    if 'last_labs' not in st.session_state:
        st.session_state.last_labs = None
    if 'last_txt_name' not in st.session_state:
        st.session_state.last_txt_name = None

    # 1. Renderiza a Sidebar e recolhe inputs
    opcoes_sidebar = render_sidebar()

    # Expander para ver os códigos dos Laboratórios Selecionados
    with st.expander("🔬 Ver Códigos (CLA) dos Laboratórios Selecionados"):
        if opcoes_sidebar['labs_selecionados']:
            st.write(
                "Os seguintes códigos estão a ser utilizados para filtrar as vendas:")
            for lab in opcoes_sidebar['labs_selecionados']:
                codigos_lab = opcoes_sidebar['dicionario_labs'].get(lab, [])
                st.markdown(f"- **{lab}:** `{', '.join(codigos_lab)}`")
        else:
            st.info("Nenhum laboratório selecionado no filtro.")

    st.divider()

    # --- Controlo Global (Aplicado a ambos os módulos) ---
    anterior = st.toggle("Média Ponderada com Base no mês ANTERIOR?")

    # Inicializa estado na sessão para manter as dataframes
    if 'df_base_agrupada' not in st.session_state:
        st.session_state.df_base_agrupada = pd.DataFrame()
        st.session_state.df_base_detalhada = pd.DataFrame()
        st.session_state.df_univ = pd.DataFrame()
        st.session_state.erros_ficheiros = []
        st.session_state.codigos_invalidos = []
        st.session_state.last_labs = None
        st.session_state.last_txt_name = None

    # ==============================================================
    # PASSO 1: Agregação Base Pesada (Ao clicar no botão)
    # ==============================================================
    if opcoes_sidebar['processar']:
        if opcoes_sidebar['ficheiros_infoprex']:
            with st.spinner("A processar e agrupar ficheiros base..."):
                # Passamos dict_locs para limpar também as localidades na base
                df_base, codigos_inv, erros = processar_ficheiros_upload(
                    ficheiros=opcoes_sidebar['ficheiros_infoprex'],
                    labs_selecionados=opcoes_sidebar['labs_selecionados'],
                    codigos_txt=opcoes_sidebar['ficheiro_codigos'],
                    _dicionario_labs=opcoes_sidebar['dicionario_labs'],
                    _dict_locs=dict_locs
                )

                if df_base.empty:
                    st.warning(
                        "Não há dados válidos para processar após aplicar os filtros.")
                    st.session_state.df_base_agrupada = pd.DataFrame()
                    st.session_state.df_base_detalhada = pd.DataFrame()
                else:
                    # 2. Criar a Tabela Master Universal de Nomes
                    df_univ = criar_tabela_dimensao(df_base)

                    # 3. Remover a DESIGNAÇÃO da base antes de agrupar para evitar conflitos textuais
                    df_base = df_base.drop(columns=['DESIGNAÇÃO'])

                    # 4. Computar as duas bases previamente (Detalhada e Agrupada)
                    df_detalhada = combina_e_agrega_df(df_base.copy(), df_univ)
                    df_detalhada = remover_linhas_sem_vendas_e_stock(
                        df_detalhada)
                    df_detalhada = df_detalhada.rename(
                        columns={'PVP': 'PVP_Médio'})

                    df_agrupada = sellout_total(df_base.copy(), df_univ)

                    # Guardar tudo em sessão
                    st.session_state.df_base_detalhada = df_detalhada
                    st.session_state.df_base_agrupada = df_agrupada
                    st.session_state.df_univ = df_univ
                    st.session_state.codigos_invalidos = codigos_inv
                    st.session_state.erros_ficheiros = erros

                    st.session_state.last_labs = opcoes_sidebar['labs_selecionados']
                    st.session_state.last_txt_name = opcoes_sidebar[
                        'ficheiro_codigos'].name if opcoes_sidebar['ficheiro_codigos'] else None
        else:
            st.sidebar.error(
                "Por favor, carregue pelo menos um ficheiro Infoprex.")

    current_txt_name = opcoes_sidebar['ficheiro_codigos'].name if opcoes_sidebar['ficheiro_codigos'] else None
    if not st.session_state.df_base_agrupada.empty and ((opcoes_sidebar['labs_selecionados'] != st.session_state.last_labs) or (current_txt_name != st.session_state.last_txt_name)):
        st.warning("⚠️ **Filtros Modificados!** Os dados apresentados abaixo encontram-se desatualizados. Clique novamente em **'Processar Dados'**.")

    for erro in st.session_state.erros_ficheiros:
        st.error(f"❌ {erro}")
    if st.session_state.codigos_invalidos:
        st.warning(
            f"⚠️ **Atenção:** As seguintes linhas foram eliminadas porque o CÓDIGO não pôde ser convertido para inteiro: {', '.join(map(str, st.session_state.codigos_invalidos))}")

    # ==============================================================
    # INTERFACE DE RESULTADOS (ABAS)
    # ==============================================================
    tab_encomendas, tab_redistribuicao = st.tabs(
        ["📊 Encomendas (Sell Out)", "🔄 Redistribuição Inteligente"])

    with tab_encomendas:
        if not st.session_state.df_base_agrupada.empty:
            # --- Contolos de UI Locais ---
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                on = st.toggle("Ver Detalhe de Sell Out?")
            with col_t2:
                st.write("")  # Espaçador

            st.subheader("Indicar o numero de meses a prever a encomenda")
            valor = st.number_input(label="Meses a prever", label_visibility="collapsed", min_value=1.0,
                                    max_value=4.0, value=1.0, step=0.1, format="%.1f", key="input_meses_encomenda")
            st.write(f"A Preparar encomenda para {valor:.1f} Meses")

            # 1. Selecionar a dataframe correta
            if on:
                df_selecionada = st.session_state.df_base_detalhada.copy()
            else:
                df_selecionada = st.session_state.df_base_agrupada.copy()

            # 2. Calcular os índices para a média ponderada
            colunas_totais = list(df_selecionada.columns)
            idx_tuni = colunas_totais.index('T Uni')
            if anterior:
                indice_colunas = [idx_tuni-2,
                                  idx_tuni-3, idx_tuni-4, idx_tuni-5]
            else:
                indice_colunas = [idx_tuni-1,
                                  idx_tuni-2, idx_tuni-3, idx_tuni-4]

            cols_selecionadas = [df_selecionada.columns[i]
                                 for i in indice_colunas]

            # 3. Processar as Propostas instantaneamente
            df_final = processar_logica_negocio(
                df_selecionada,
                df_esgotados,
                df_nao_comprar,
                cols_selecionadas,
                [0.4, 0.3, 0.2, 0.1],
                valor,
                agrupado=not on
            )

            st.success(f"Dados prontos! Total de linhas: {len(df_final)}")

            # 4. Renderização Web (Ocultando o CLA)
            df_view = df_final.drop(columns=['CLA'], errors='ignore')
            st_styled = df_view.style.apply(aplicar_destaques, axis=1).format(
                {'PVP_Médio': '{:.2f}', 'P.CUSTO_Médio': '{:.2f}', 'P.CUSTO': '{:.2f}'}, na_rep="")
            st.dataframe(st_styled, width='stretch', hide_index=True)

            # 5. Exportação Excel
            excel_data = formatar_excel(df_view)
            st.download_button(
                label="Download Excel Encomendas",
                data=excel_data,
                file_name='Sell_Out_GRUPO.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            st.info(
                "Nenhum dado processado ainda. Configure a sidebar e clique em 'Processar Dados'.")

    with tab_redistribuicao:
        if not st.session_state.df_base_detalhada.empty:
            st.error("### 🚨 **ATENÇÃO: FUNÇÃO EM DESENVOLVIMENTO** 🚨")
            st.warning(
                "### ⚠️ **Aviso:** Esta função ainda está em desenvolvimento | Tem erros de calculo | Apenas para Desenvolvimento. ⚠️")
            st.subheader("⚙️ Configurações de Redistribuição")
            col_r1, col_r2 = st.columns(2)
            with col_r1:
                meses_val = st.slider(
                    "Meses mínimos de Validade (Não transferir se caducar antes)", 1, 12, 3)
            with col_r2:
                dias_imun = st.slider(
                    "Dias de Imunidade para Produto Novo", 0, 180, 90)

            if st.button("🚀 Gerar Plano de Redistribuição", type="primary", width='stretch'):
                with st.spinner("A gerar transferências inteligentes..."):
                    df_base_reorder = st.session_state.df_base_detalhada.copy()

                    colunas_totais = list(df_base_reorder.columns)
                    idx_tuni = colunas_totais.index('T Uni')
                    if anterior:
                        indice_colunas = [idx_tuni-2,
                                          idx_tuni-3, idx_tuni-4, idx_tuni-5]
                    else:
                        indice_colunas = [idx_tuni-1,
                                          idx_tuni-2, idx_tuni-3, idx_tuni-4]

                    cols_vendas = [df_base_reorder.columns[i]
                                   for i in indice_colunas]

                    plano_df = gerar_plano_redistribuicao(
                        df_base_reorder,
                        st.session_state.df_univ,
                        meses_val,
                        dias_imun,
                        cols_vendas,
                        [0.4, 0.3, 0.2, 0.1]
                    )

                    if not plano_df.empty:
                        st.success(
                            f"Plano gerado com sucesso! {len(plano_df)} transferências sugeridas.")
                        st.dataframe(plano_df, width='stretch',
                                     hide_index=True)

                        # Output Excel do Plano
                        output_plan = BytesIO()
                        plano_df.to_excel(output_plan, index=False)
                        output_plan.seek(0)
                        st.download_button(
                            label="Download Plano de Redistribuição",
                            data=output_plan,
                            file_name='Plano_Redistribuicao.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                    else:
                        st.info(
                            "O algoritmo não encontrou nenhuma oportunidade de transferência segura para este portefólio.")
        else:
            st.info(
                "Nenhum dado processado ainda. Configure a sidebar e clique em 'Processar Dados'.")


if __name__ == "__main__":
    main()
